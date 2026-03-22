"""
Build the IC input workbook with demo data.
5 entity IC Detail tabs + 1 IC Agreement Schedule + 1 Entity Registry.
Contains 7 transaction sets with 5 planted errors.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from decimal import Decimal

wb = openpyxl.Workbook()

# --- Styles ---
header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Arial", size=10)
error_fill = PatternFill("solid", fgColor="FFF2CC")  # Subtle yellow — planted errors
title_font = Font(name="Arial", bold=True, size=12, color="2F5496")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
acct_fmt = '0000'
usd_fmt = '#,##0.00'
rate_fmt = '0.0000'
local_fmt = '#,##0.00'

IC_DETAIL_HEADERS = [
    "Transaction ID",
    "Posting Date",
    "Entity ID",
    "IC Partner Entity",
    "IC Account Code",
    "Account Name",
    "Transaction Type",
    "Description",
    "Local Currency",
    "Local Amount",
    "FX Rate (to USD)",
    "USD Equivalent",
    "Settlement Status",
    "Reference Number",
    "Explanation",
]

def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = data_font
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    return cell

def auto_width(ws, num_cols, min_width=12, max_width=35):
    for col in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(min_width, 15), max_width)
    ws.column_dimensions[get_column_letter(num_cols)].width = 50  # Explanation col

def add_ic_detail_rows(ws, rows, error_rows=None):
    error_rows = error_rows or []
    for i, row_data in enumerate(rows, start=2):
        for j, val in enumerate(row_data, start=1):
            cell = style_data_cell(ws, i, j)
            cell.value = val
            if j == 5:  # Account code
                cell.number_format = acct_fmt
            elif j in (10,):  # Local amount
                cell.number_format = local_fmt
            elif j == 11:  # FX rate
                cell.number_format = rate_fmt
            elif j == 12:  # USD equivalent
                cell.number_format = usd_fmt
        if i in error_rows:
            for j in range(1, len(row_data) + 1):
                ws.cell(row=i, column=j).fill = error_fill


# ============================================================
# TAB 1: APX-TECH (Entity 1 — Apex Technologies Inc.)
# ============================================================
ws = wb.active
ws.title = "APX-TECH IC Detail"
ws.append(IC_DETAIL_HEADERS)
style_header_row(ws, len(IC_DETAIL_HEADERS))

apx_tech_data = [
    # Transaction Set 1: Clean trade sale to APX-UK ($2,400,000)
    [
        "TECH-001", "2025-01-15", "APX-TECH", "APX-UK", "4100",
        "IC Revenue - Product Sales", "Trade",
        "Q1 product license sale to Apex UK - enterprise security suite",
        "USD", 2400000.00, 1.0000, 2400000.00, "Open", "REF-TS1-001",
        "Quarterly enterprise security license sale per master IC agreement. Standard pricing applied.",
    ],
    [
        "TECH-002", "2025-01-15", "APX-TECH", "APX-UK", "1310",
        "IC Receivable - Trade", "Trade",
        "Receivable from Apex UK for Q1 product license sale",
        "USD", 2400000.00, 1.0000, 2400000.00, "Open", "REF-TS1-001",
        "Receivable recognized at point of sale. 60-day payment terms.",
    ],
    # Transaction Set 5: FX rate mismatch — PLANTED ERROR #2
    # Entity 1 used transaction date rate 1.2750 to derive $1,850,000
    # The GBP amount implied: 1,850,000 / 1.2750 = 1,450,980.39
    [
        "TECH-003", "2025-02-20", "APX-TECH", "APX-UK", "4100",
        "IC Revenue - Product Sales", "Trade",
        "Q1 cloud migration services provided to Apex UK",
        "USD", 1850000.00, 1.0000, 1850000.00, "Open", "REF-TS5-001",
        "Cloud migration consulting engagement. Transaction date rate of 1.2750 GBP/USD used per policy.",
    ],
    [
        "TECH-004", "2025-02-20", "APX-TECH", "APX-UK", "1310",
        "IC Receivable - Trade", "Trade",
        "Receivable from Apex UK for cloud migration services",
        "USD", 1850000.00, 1.0000, 1850000.00, "Open", "REF-TS5-001",
        "Receivable at transaction date rate. Standard 60-day terms.",
    ],
    # Transaction Set 7: Clean dividend from APX-UK ($500,000)
    [
        "TECH-005", "2025-03-15", "APX-TECH", "APX-UK", "1340",
        "IC Receivable - Dividends", "Dividend",
        "Q1 dividend declared by Apex UK Ltd",
        "USD", 500000.00, 1.0000, 500000.00, "Open", "REF-TS7-001",
        "Annual dividend per board resolution dated March 1, 2025.",
    ],
    [
        "TECH-006", "2025-03-15", "APX-TECH", "APX-UK", "3520",
        "IC Equity - Dividend Received", "Dividend",
        "Dividend income from Apex UK subsidiary",
        "USD", 500000.00, 1.0000, 500000.00, "Open", "REF-TS7-001",
        "Recorded per ASC 810. Dividend receipt confirmed by UK board minutes.",
    ],
    # Transaction Set 3: Mgmt fee from APX-SS — PLANTED ERROR #4
    # Entity 1 posted to WRONG account: 6200 (IT Allocation) instead of 6100 (Mgmt Fee)
    [
        "TECH-007", "2025-03-31", "APX-TECH", "APX-SS", "6200",
        "IC Expense - IT Allocation", "MgmtFee",
        "Q1 management fee from Apex Shared Services",
        "USD", 250000.00, 1.0000, 250000.00, "Open", "REF-TS3-TECH",
        "Quarterly shared services allocation. Charged based on headcount methodology.",
    ],
    [
        "TECH-008", "2025-03-31", "APX-TECH", "APX-SS", "2330",
        "IC Payable - Mgmt Fee", "MgmtFee",
        "Payable to Apex Shared Services for Q1 management fee",
        "USD", 250000.00, 1.0000, 250000.00, "Open", "REF-TS3-TECH",
        "Payable recognized per IC agreement. Net 30 settlement.",
    ],
]

add_ic_detail_rows(ws, apx_tech_data, error_rows=[9, 10])  # rows for TECH-007/008 = error #4
auto_width(ws, len(IC_DETAIL_HEADERS))
ws.freeze_panes = "A2"


# ============================================================
# TAB 2: APX-CS (Entity 2 — Apex Cloud Services LLC)
# ============================================================
ws2 = wb.create_sheet("APX-CS IC Detail")
ws2.append(IC_DETAIL_HEADERS)
style_header_row(ws2, len(IC_DETAIL_HEADERS))

apx_cs_data = [
    # Transaction Set 2: IC Loan to APX-DE — PLANTED ERROR #3
    # Entity 2 says remaining balance is $1,200,000 (received $300K repayment)
    [
        "CS-001", "2025-01-10", "APX-CS", "APX-DE", "1320",
        "IC Receivable - Loan", "Loan",
        "IC loan to Apex Deutschland - original $1,500,000, partial repayment received",
        "USD", 1200000.00, 1.0000, 1200000.00, "Partial", "REF-TS2-001",
        "Original loan of $1,500,000 funded Jan 2024. Received partial repayment of $300,000 on Feb 15, 2025. Remaining balance $1,200,000.",
    ],
    # Transaction Set 4: One-sided posting — PLANTED ERROR #1
    # Entity 2 booked revenue + receivable, but Entity 4 booked NOTHING
    [
        "CS-002", "2025-03-20", "APX-CS", "APX-DE", "4100",
        "IC Revenue - Product Sales", "Trade",
        "Q1 cloud infrastructure services provided to Apex Deutschland",
        "USD", 475000.00, 1.0000, 475000.00, "Open", "REF-TS4-001",
        "Cloud infrastructure hosting services for Q1. Invoice sent March 20, 2025.",
    ],
    [
        "CS-003", "2025-03-20", "APX-CS", "APX-DE", "1310",
        "IC Receivable - Trade", "Trade",
        "Receivable from Apex Deutschland for cloud infrastructure services",
        "USD", 475000.00, 1.0000, 475000.00, "Open", "REF-TS4-001",
        "Receivable recognized per completed service delivery. 45-day payment terms.",
    ],
    # Transaction Set 3: Mgmt fee from APX-SS (clean)
    [
        "CS-004", "2025-03-31", "APX-CS", "APX-SS", "6100",
        "IC Expense - Mgmt Fee", "MgmtFee",
        "Q1 management fee from Apex Shared Services",
        "USD", 200000.00, 1.0000, 200000.00, "Open", "REF-TS3-CS",
        "Quarterly shared services allocation per IC agreement.",
    ],
    [
        "CS-005", "2025-03-31", "APX-CS", "APX-SS", "2330",
        "IC Payable - Mgmt Fee", "MgmtFee",
        "Payable to Apex Shared Services for Q1 management fee",
        "USD", 200000.00, 1.0000, 200000.00, "Open", "REF-TS3-CS",
        "Payable per monthly allocation schedule.",
    ],
]

add_ic_detail_rows(ws2, apx_cs_data, error_rows=[2, 3, 4])  # CS-001=error#3, CS-002/003=error#1
auto_width(ws2, len(IC_DETAIL_HEADERS))
ws2.freeze_panes = "A2"


# ============================================================
# TAB 3: APX-UK (Entity 3 — Apex UK Ltd)
# ============================================================
ws3 = wb.create_sheet("APX-UK IC Detail")
ws3.append(IC_DETAIL_HEADERS)
style_header_row(ws3, len(IC_DETAIL_HEADERS))

apx_uk_data = [
    # Transaction Set 1: Clean trade purchase from APX-TECH
    [
        "UK-001", "2025-01-15", "APX-UK", "APX-TECH", "5100",
        "IC COGS - Product Purchases", "Trade",
        "Q1 enterprise security suite purchase from Apex Technologies",
        "GBP", 1882352.94, 1.2750, 2400000.00, "Open", "REF-TS1-001",
        "Purchase of enterprise security license. Rate per Treasury: 1.2750 GBP/USD.",
    ],
    [
        "UK-002", "2025-01-15", "APX-UK", "APX-TECH", "2310",
        "IC Payable - Trade", "Trade",
        "Payable to Apex Technologies for enterprise security suite",
        "GBP", 1882352.94, 1.2750, 2400000.00, "Open", "REF-TS1-001",
        "Payable recorded at transaction date spot rate.",
    ],
    # Transaction Set 5: FX rate mismatch — PLANTED ERROR #2
    # Entity 3 used MONTH-END rate 1.2600 instead of transaction date rate 1.2750
    # GBP amount: 1,850,000 / 1.2600 = 1,468,253.97
    # USD equivalent at 1.2600: 1,468,253.97 * 1.2600 = 1,850,000.00
    # The USD amounts MATCH but the rates diverge
    [
        "UK-003", "2025-02-20", "APX-UK", "APX-TECH", "5100",
        "IC COGS - Product Purchases", "Trade",
        "Q1 cloud migration services received from Apex Technologies",
        "GBP", 1468253.97, 1.2600, 1850000.00, "Open", "REF-TS5-001",
        "Cloud migration engagement. Month-end rate of 1.2600 applied per local accounting policy.",
    ],
    [
        "UK-004", "2025-02-20", "APX-UK", "APX-TECH", "2310",
        "IC Payable - Trade", "Trade",
        "Payable to Apex Technologies for cloud migration services",
        "GBP", 1468253.97, 1.2600, 1850000.00, "Open", "REF-TS5-001",
        "Payable at month-end rate per UK entity policy. Rate confirmed by local Treasury.",
    ],
    # Transaction Set 7: Clean dividend to APX-TECH
    [
        "UK-005", "2025-03-15", "APX-UK", "APX-TECH", "3510",
        "IC Equity - Dividend Paid", "Dividend",
        "Q1 dividend declared to Apex Technologies parent",
        "GBP", 392156.86, 1.2750, 500000.00, "Open", "REF-TS7-001",
        "Dividend per board resolution. Rate per Treasury at declaration date.",
    ],
    [
        "UK-006", "2025-03-15", "APX-UK", "APX-TECH", "2340",
        "IC Payable - Dividends", "Dividend",
        "Dividend payable to Apex Technologies parent",
        "GBP", 392156.86, 1.2750, 500000.00, "Open", "REF-TS7-001",
        "Payable to parent entity. Funds transfer scheduled for April 15, 2025.",
    ],
    # Transaction Set 6: Unauthorized IC pair — PLANTED ERROR #5
    # APX-UK <-> APX-DE is NOT on the IC Agreement Schedule
    [
        "UK-007", "2025-03-25", "APX-UK", "APX-DE", "4100",
        "IC Revenue - Product Sales", "Trade",
        "Direct sale of cybersecurity module to Apex Deutschland",
        "GBP", 120000.00, 1.2750, 153000.00, "Open", "REF-TS6-001",
        "One-time sale of cybersecurity module. Approved by UK regional VP.",
    ],
    [
        "UK-008", "2025-03-25", "APX-UK", "APX-DE", "1310",
        "IC Receivable - Trade", "Trade",
        "Receivable from Apex Deutschland for cybersecurity module",
        "GBP", 120000.00, 1.2750, 153000.00, "Open", "REF-TS6-001",
        "Receivable per standard terms. 30-day payment.",
    ],
    # Transaction Set 3: Mgmt fee from APX-SS (clean)
    [
        "UK-009", "2025-03-31", "APX-UK", "APX-SS", "6100",
        "IC Expense - Mgmt Fee", "MgmtFee",
        "Q1 management fee from Apex Shared Services",
        "GBP", 156862.75, 1.2750, 200000.00, "Open", "REF-TS3-UK",
        "Quarterly shared services allocation. Converted at transaction date rate.",
    ],
    [
        "UK-010", "2025-03-31", "APX-UK", "APX-SS", "2330",
        "IC Payable - Mgmt Fee", "MgmtFee",
        "Payable to Apex Shared Services for Q1 management fee",
        "GBP", 156862.75, 1.2750, 200000.00, "Open", "REF-TS3-UK",
        "Payable recognized per IC agreement terms.",
    ],
]

add_ic_detail_rows(ws3, apx_uk_data, error_rows=[4, 5, 8, 9])  # error#2 and error#5
auto_width(ws3, len(IC_DETAIL_HEADERS))
ws3.freeze_panes = "A2"


# ============================================================
# TAB 4: APX-DE (Entity 4 — Apex Deutschland GmbH)
# ============================================================
ws4 = wb.create_sheet("APX-DE IC Detail")
ws4.append(IC_DETAIL_HEADERS)
style_header_row(ws4, len(IC_DETAIL_HEADERS))

apx_de_data = [
    # Transaction Set 2: IC Loan from APX-CS — PLANTED ERROR #3
    # Entity 4 says remaining balance is $1,250,000 (paid $250K repayment)
    # Mismatch: Entity 2 says $1,200,000 remaining (received $300K)
    [
        "DE-001", "2025-01-10", "APX-DE", "APX-CS", "2320",
        "IC Payable - Loan", "Loan",
        "IC loan from Apex Cloud Services - original $1,500,000, partial repayment made",
        "EUR", 1086956.52, 1.1500, 1250000.00, "Partial", "REF-TS2-001",
        "Original loan $1,500,000 received Jan 2024. Made partial repayment of $250,000 on Feb 15, 2025. Remaining balance $1,250,000.",
    ],
    # Transaction Set 4: One-sided — PLANTED ERROR #1
    # Entity 4 has NO entries for the $475,000 trade from APX-CS
    # (This is shown by the ABSENCE of rows, which the matching engine detects)

    # Transaction Set 6: Unauthorized IC pair — PLANTED ERROR #5
    # APX-DE <-> APX-UK is NOT authorized
    [
        "DE-002", "2025-03-25", "APX-DE", "APX-UK", "5100",
        "IC COGS - Product Purchases", "Trade",
        "Purchase of cybersecurity module from Apex UK",
        "EUR", 133043.48, 1.1500, 153000.00, "Open", "REF-TS6-001",
        "Cybersecurity module purchase. Approved by DE operations manager.",
    ],
    [
        "DE-003", "2025-03-25", "APX-DE", "APX-UK", "2310",
        "IC Payable - Trade", "Trade",
        "Payable to Apex UK for cybersecurity module",
        "EUR", 133043.48, 1.1500, 153000.00, "Open", "REF-TS6-001",
        "Payable per standard terms.",
    ],
    # Transaction Set 3: Mgmt fee from APX-SS (clean)
    [
        "DE-004", "2025-03-31", "APX-DE", "APX-SS", "6100",
        "IC Expense - Mgmt Fee", "MgmtFee",
        "Q1 management fee from Apex Shared Services",
        "EUR", 130434.78, 1.1500, 150000.00, "Open", "REF-TS3-DE",
        "Quarterly shared services allocation per IC agreement.",
    ],
    [
        "DE-005", "2025-03-31", "APX-DE", "APX-SS", "2330",
        "IC Payable - Mgmt Fee", "MgmtFee",
        "Payable to Apex Shared Services for Q1 management fee",
        "EUR", 130434.78, 1.1500, 150000.00, "Open", "REF-TS3-DE",
        "Payable recognized. Net 30 terms.",
    ],
]

add_ic_detail_rows(ws4, apx_de_data, error_rows=[2, 3, 4])  # error#3 and error#5
auto_width(ws4, len(IC_DETAIL_HEADERS))
ws4.freeze_panes = "A2"


# ============================================================
# TAB 5: APX-SS (Entity 5 — Apex Shared Services Corp)
# ============================================================
ws5 = wb.create_sheet("APX-SS IC Detail")
ws5.append(IC_DETAIL_HEADERS)
style_header_row(ws5, len(IC_DETAIL_HEADERS))

apx_ss_data = [
    # Transaction Set 3: Management fees charged to all entities
    # Revenue side (charger)
    [
        "SS-001", "2025-03-31", "APX-SS", "APX-TECH", "4200",
        "IC Revenue - Services", "MgmtFee",
        "Q1 management fee charged to Apex Technologies",
        "USD", 250000.00, 1.0000, 250000.00, "Open", "REF-TS3-TECH",
        "Quarterly allocation based on headcount. APX-TECH: 250 FTEs of 1,000 total = 25%.",
    ],
    [
        "SS-002", "2025-03-31", "APX-SS", "APX-TECH", "1330",
        "IC Receivable - Mgmt Fee", "MgmtFee",
        "Receivable from Apex Technologies for Q1 management fee",
        "USD", 250000.00, 1.0000, 250000.00, "Open", "REF-TS3-TECH",
        "Receivable per allocation schedule. Net 30.",
    ],
    [
        "SS-003", "2025-03-31", "APX-SS", "APX-CS", "4200",
        "IC Revenue - Services", "MgmtFee",
        "Q1 management fee charged to Apex Cloud Services",
        "USD", 200000.00, 1.0000, 200000.00, "Open", "REF-TS3-CS",
        "Quarterly allocation. APX-CS: 200 FTEs = 20%.",
    ],
    [
        "SS-004", "2025-03-31", "APX-SS", "APX-CS", "1330",
        "IC Receivable - Mgmt Fee", "MgmtFee",
        "Receivable from Apex Cloud Services for Q1 management fee",
        "USD", 200000.00, 1.0000, 200000.00, "Open", "REF-TS3-CS",
        "Receivable per allocation schedule.",
    ],
    [
        "SS-005", "2025-03-31", "APX-SS", "APX-UK", "4200",
        "IC Revenue - Services", "MgmtFee",
        "Q1 management fee charged to Apex UK",
        "USD", 200000.00, 1.0000, 200000.00, "Open", "REF-TS3-UK",
        "Quarterly allocation. APX-UK: 200 FTEs = 20%.",
    ],
    [
        "SS-006", "2025-03-31", "APX-SS", "APX-UK", "1330",
        "IC Receivable - Mgmt Fee", "MgmtFee",
        "Receivable from Apex UK for Q1 management fee",
        "USD", 200000.00, 1.0000, 200000.00, "Open", "REF-TS3-UK",
        "Receivable per allocation schedule.",
    ],
    [
        "SS-007", "2025-03-31", "APX-SS", "APX-DE", "4200",
        "IC Revenue - Services", "MgmtFee",
        "Q1 management fee charged to Apex Deutschland",
        "USD", 150000.00, 1.0000, 150000.00, "Open", "REF-TS3-DE",
        "Quarterly allocation. APX-DE: 150 FTEs = 15%.",
    ],
    [
        "SS-008", "2025-03-31", "APX-SS", "APX-DE", "1330",
        "IC Receivable - Mgmt Fee", "MgmtFee",
        "Receivable from Apex Deutschland for Q1 management fee",
        "USD", 150000.00, 1.0000, 150000.00, "Open", "REF-TS3-DE",
        "Receivable per allocation schedule.",
    ],
]

add_ic_detail_rows(ws5, apx_ss_data)
auto_width(ws5, len(IC_DETAIL_HEADERS))
ws5.freeze_panes = "A2"


# ============================================================
# TAB 6: IC Agreement Schedule
# ============================================================
ws6 = wb.create_sheet("IC Agreement Schedule")
agree_headers = [
    "Entity A", "Entity A Name",
    "Entity B", "Entity B Name",
    "Authorized IC Types", "Status", "Effective Date", "Notes",
]
ws6.append(agree_headers)
style_header_row(ws6, len(agree_headers))

agreements = [
    ["APX-TECH", "Apex Technologies Inc.", "APX-UK", "Apex UK Ltd",
     "Trade, Dividend", "Active", "2020-01-01", "Parent-sub relationship. Trade includes product licenses and services."],
    ["APX-TECH", "Apex Technologies Inc.", "APX-SS", "Apex Shared Services Corp",
     "MgmtFee", "Active", "2021-07-01", "Headcount-based quarterly allocation."],
    ["APX-CS", "Apex Cloud Services LLC", "APX-DE", "Apex Deutschland GmbH",
     "Trade, Loan", "Active", "2022-01-01", "Cloud services and intercompany lending facility."],
    ["APX-CS", "Apex Cloud Services LLC", "APX-SS", "Apex Shared Services Corp",
     "MgmtFee", "Active", "2021-07-01", "Headcount-based quarterly allocation."],
    ["APX-UK", "Apex UK Ltd", "APX-SS", "Apex Shared Services Corp",
     "MgmtFee", "Active", "2021-07-01", "Headcount-based quarterly allocation."],
    ["APX-DE", "Apex Deutschland GmbH", "APX-SS", "Apex Shared Services Corp",
     "MgmtFee", "Active", "2021-07-01", "Headcount-based quarterly allocation."],
]

for i, row in enumerate(agreements, start=2):
    for j, val in enumerate(row, start=1):
        cell = style_data_cell(ws6, i, j)
        cell.value = val

for col in range(1, len(agree_headers) + 1):
    ws6.column_dimensions[get_column_letter(col)].width = 22
ws6.column_dimensions["H"].width = 55
ws6.freeze_panes = "A2"


# ============================================================
# TAB 7: Entity Registry
# ============================================================
ws7 = wb.create_sheet("Entity Registry")
entity_headers = [
    "Entity ID", "Entity Name", "Functional Currency",
    "Country", "Segment", "Status",
]
ws7.append(entity_headers)
style_header_row(ws7, len(entity_headers))

entities = [
    ["APX-TECH", "Apex Technologies Inc.", "USD", "US", "Enterprise Solutions", "Active"],
    ["APX-CS", "Apex Cloud Services LLC", "USD", "US", "Cloud & Infrastructure", "Active"],
    ["APX-UK", "Apex UK Ltd", "GBP", "UK", "Enterprise Solutions", "Active"],
    ["APX-DE", "Apex Deutschland GmbH", "EUR", "DE", "Cloud & Infrastructure", "Active"],
    ["APX-SS", "Apex Shared Services Corp", "USD", "US", "Corporate", "Active"],
]

for i, row in enumerate(entities, start=2):
    for j, val in enumerate(row, start=1):
        cell = style_data_cell(ws7, i, j)
        cell.value = val

for col in range(1, len(entity_headers) + 1):
    ws7.column_dimensions[get_column_letter(col)].width = 25
ws7.freeze_panes = "A2"


# ============================================================
# Save
# ============================================================
output_path = "/home/claude/ic_elimination_agent/ic_input_workbook.xlsx"
wb.save(output_path)
print(f"Input workbook saved to {output_path}")
print(f"Tabs: {wb.sheetnames}")
print(f"Total IC transactions: {len(apx_tech_data) + len(apx_cs_data) + len(apx_uk_data) + len(apx_de_data) + len(apx_ss_data)}")
