"""
Reads the IC input workbook and returns validated ICTransaction objects
plus the IC Agreement Schedule.
"""

from decimal import Decimal, InvalidOperation
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from models import ICTransaction
import config


def parse_workbook(filepath: str) -> Tuple[List[ICTransaction], Dict[frozenset, List[str]]]:
    wb = load_workbook(filepath, data_only=True)
    transactions = []
    errors = []

    # Parse entity IC Detail tabs
    for entity_id, entity_info in config.ENTITIES.items():
        sheet_name = f"{entity_id} IC Detail"
        if sheet_name not in wb.sheetnames:
            errors.append(f"Missing tab: {sheet_name}")
            continue
        ws = wb[sheet_name]
        entity_txns, entity_errors = _parse_ic_detail_tab(ws, entity_id)
        transactions.extend(entity_txns)
        errors.extend(entity_errors)

    # Parse IC Agreement Schedule
    agreement_schedule = _parse_agreement_schedule(wb)

    if errors:
        print(f"\n  VALIDATION WARNINGS ({len(errors)}):")
        for e in errors:
            print(f"    - {e}")

    print(f"  Parsed {len(transactions)} IC transactions across {len(config.ENTITIES)} entities")
    return transactions, agreement_schedule


def _parse_ic_detail_tab(ws, entity_id: str) -> Tuple[List[ICTransaction], List[str]]:
    transactions = []
    errors = []
    headers = [cell.value for cell in ws[1]]

    col_map = {}
    expected_cols = [
        "Transaction ID", "Posting Date", "Entity ID", "IC Partner Entity",
        "IC Account Code", "Account Name", "Transaction Type", "Description",
        "Local Currency", "Local Amount", "FX Rate (to USD)", "USD Equivalent",
        "Settlement Status", "Reference Number", "Explanation",
    ]
    for col_name in expected_cols:
        if col_name in headers:
            col_map[col_name] = headers.index(col_name)
        else:
            errors.append(f"{entity_id}: Missing column '{col_name}'")

    if len(errors) > 0:
        return transactions, errors

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue

        try:
            txn_id = str(row[col_map["Transaction ID"]])
            partner = str(row[col_map["IC Partner Entity"]])

            if partner not in config.ENTITIES:
                errors.append(f"{entity_id} row {row_idx}: Unknown partner entity '{partner}'")

            acct_code = str(row[col_map["IC Account Code"]]).zfill(4)
            if acct_code not in config.CHART_OF_ACCOUNTS:
                errors.append(f"{entity_id} row {row_idx}: Unknown account code '{acct_code}'")

            local_amt = _to_decimal(row[col_map["Local Amount"]])
            fx_rate = _to_decimal_rate(row[col_map["FX Rate (to USD)"]])
            usd_amt = _to_decimal(row[col_map["USD Equivalent"]])

            # Validate FX math (local * rate ≈ USD, within tolerance)
            if fx_rate > 0:
                expected_usd = local_amt * fx_rate
                if abs(expected_usd - usd_amt) > Decimal("100.00"):
                    errors.append(
                        f"{entity_id} {txn_id}: FX math doesn't tie. "
                        f"{local_amt} * {fx_rate} = {expected_usd}, but USD shows {usd_amt}"
                    )

            txn = ICTransaction(
                transaction_id=txn_id,
                entity_id=str(row[col_map["Entity ID"]]),
                partner_entity_id=partner,
                account_code=acct_code,
                account_name=str(row[col_map["Account Name"]] or ""),
                transaction_type=str(row[col_map["Transaction Type"]]),
                description=str(row[col_map["Description"]] or ""),
                local_currency=str(row[col_map["Local Currency"]]),
                local_amount=local_amt,
                fx_rate=fx_rate,
                usd_amount=usd_amt,
                settlement_status=str(row[col_map["Settlement Status"]] or "Open"),
                reference_number=str(row[col_map["Reference Number"]] or ""),
                posting_date=str(row[col_map["Posting Date"]] or ""),
                explanation=str(row[col_map["Explanation"]] or ""),
            )
            transactions.append(txn)

        except Exception as e:
            errors.append(f"{entity_id} row {row_idx}: Parse error — {e}")

    return transactions, errors


def _parse_agreement_schedule(wb) -> Dict[frozenset, List[str]]:
    if "IC Agreement Schedule" not in wb.sheetnames:
        print("  WARNING: No IC Agreement Schedule tab found. Using config defaults.")
        return config.AUTHORIZED_IC_PAIRS

    ws = wb["IC Agreement Schedule"]
    schedule = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        entity_a = str(row[0])
        entity_b = str(row[2])
        ic_types = [t.strip() for t in str(row[4]).split(",")]
        pair_key = frozenset({entity_a, entity_b})
        schedule[pair_key] = ic_types
    return schedule


def _to_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_decimal_rate(val) -> Decimal:
    """Convert FX rates preserving 4 decimal places."""
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        return Decimal("0")
